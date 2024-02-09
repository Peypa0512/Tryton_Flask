from flask import Flask, render_template, request, send_file
from flask_tryton import Tryton
import json
from flask_cors import CORS
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side



app = Flask(__name__)
CORS(app)
app.config['TRYTON_DATABASE'] = 'prueba'
tryton = Tryton(app, configure_jinja=True)
Prueba = tryton.pool.get('ack.prueba')

@app.route('/export', methods=['POST'])
@tryton.transaction()
def export():

    if request.method == 'POST':
        data_checkbox = request.form.getlist('checkboxes')
        past = []
        # recoger datos
        for fact in data_checkbox:
            value_catch = Prueba.search([('id', '=', fact)])
            if value_catch:
                past.append(value_catch)

        return render_template('excelExport.html', values=past)
    return render_template('hecho')


@app.route("/export2", methods=['POST'])
def export2():
    selected_ids = request.form.getlist('checkboxes')

    # Obtén las filas seleccionadas directamente del formulario
    selected_lines = []
    for line_id in selected_ids:
        fecha = request.form.get(f'line_date_{line_id}')
        dedicacion = request.form.get(f'line_duration_{line_id}')
        tipo_tarea = request.form.get(f'line_work-type_{line_id}')
        estado_tarea = request.form.get(f'line_status_{line_id}')
        tecnico = request.form.get(f'line_employee_{line_id}')
        cliente = request.form.get(f'line_party_{line_id}')
        tarea = request.form.get(f'line_work_{line_id}')
        descripcion = request.form.get(f'line_description_{line_id}')
        editor_ver = request.form.get(f'line_internal-description_{line_id}')

        # Agrega los datos de la fila a la lista
        selected_lines.append({
            'fecha': fecha,
            'dedicacion': dedicacion,
            'tipo_tarea': tipo_tarea,
            'estado_tarea': estado_tarea,
            'tecnico': tecnico,
            'cliente': cliente,
            'tarea': tarea,
            'descripcion': descripcion,
            'editor_ver': editor_ver
        })

    return render_template('excelExport.html', lines=selected_lines)


@app.route('/excelExport', methods=['POST'])
def excelExport():
    print('ok')
    # Obtener los datos del campo oculto
    # Obtener los datos del campo oculto
    data_json = request.form.get('data')
    def agregar_tildes(cadena):
        tildes = {'dedicacion': 'Dedicación', 'tecnico': 'Técnico',
                  'descripcion': 'Descripción'}  # Agrega más palabras según sea necesario
        words = cadena.split()
        result = ' '.join(tildes.get(word.lower(), word) for word in words)
        return result

    try:

        # Cargar datos JSON
        data = json.loads(data_json)
    except json.decoder.JSONDecodeError as e:
        print("Error al decodificar JSON:", str(e))
        return "Error al decodificar JSON"

    dataframes = []
    diccionary_comb = {}
    # Asegúrate de que hay datos y al menos un diccionario en la lista
    if data and isinstance(data[0], list):
        # Iterar sobre cada conjunto de datos en la lista
        for conjunto_data in data:
            # Crear un diccionario combinado para el conjunto actual

            for diccionary in conjunto_data:
                diccionary_comb.update(diccionary)

            # Crear DataFrame de Pandas con el conjunto actual de datos
            df = pd.DataFrame([diccionary_comb])

            # Aplicar la función agregar_tildes al nombrar las columnas del DataFrame
            df.columns = df.columns.map(agregar_tildes)

            # Agregar el DataFrame actual a la lista
            dataframes.append(df)
        print('llegará?')
        # Concatenar todos los DataFrames en uno solo
        df_final = pd.concat(dataframes, ignore_index=True)

        # Reemplazar caracteres con tilde en las cabeceras
        df_final = df_final.rename(columns=lambda x: agregar_tildes(x))

        # Crear un nuevo libro de trabajo de Excel con openpyxl
        book = Workbook()

        # Seleccionar la hoja de trabajo activa
        sheet = book.active

        # Configurar el estilo para el encabezado (negrita y centrado)
        header_style= Font(bold=True)
        center_align= Alignment(horizontal='center', vertical='center')

        # Configurar el estilo para el color de fondo de la cabecera
        fill_header = PatternFill(start_color="223776", end_color="223776", fill_type="solid")

        # Configurar el estilo para el color de fondo de la tabla
        fill_table = PatternFill(start_color="b3d7ff", end_color="b3d7ff", fill_type="solid")

        # Después de insertar los encabezados en la hoja de trabajo
        for columns, header in enumerate(df_final.columns, start=1):
            cell = sheet.cell(row=1, column=columns, value=header)
            cell.font = header_style
            cell.alignment = center_align

            # Ajustar el ancho de la columna según la longitud máxima del encabezado o los datos en esa columna
            long_max= max(len(str(header)), df_final[header].astype(str).apply(len).max()) + 2
            sheet.column_dimensions[sheet.cell(row=1, column=columns).column_letter].width = long_max

            cell.fill = fill_header  # Color de fondo para la cabecera

            # Configurar bordes para cada celda
            borders = Border(left=Side(border_style='thin'), right=Side(border_style='thin'),
                            top=Side(border_style='thin'), bottom=Side(border_style='thin'))
            cell.border = borders

        # Después de insertar los datos en la hoja de trabajo
        for file, values in enumerate(df_final.values, start=2):
            for columns, value in enumerate(values, start=1):
                cell = sheet.cell(row=file, column=columns, value=value)
                cell.fill = fill_table  # Color de fondo para la tabla

                # Configurar bordes para cada celda
                borders = Border(left=Side(border_style='thin'), right=Side(border_style='thin'),
                                top=Side(border_style='thin'), bottom=Side(border_style='thin'))
                cell.border = borders

                # Justificar a la derecha
                cell.alignment = Alignment(horizontal='right')
        print('todo OK, antes de guardar el fichero, tiene que pintar todo')

        # Guardar el libro de trabajo en un archivo Excel
        book.save('excel_export.xlsx')

        # Devolver el archivo Excel como respuesta
        return send_file('excel_export.xlsx', as_attachment=True)
    else:
        return "Datos no válidos"
if __name__ == '__main__':
    app.run(debug=True)
