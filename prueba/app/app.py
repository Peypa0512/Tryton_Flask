from flask import Flask, render_template, request, send_file, jsonify, json, Response, url_for, redirect
from flask_tryton import Tryton
import json
from flask_cors import CORS
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import pdb
from datetime import datetime, timedelta


app = Flask(__name__)
CORS(app)
app.config['TRYTON_DATABASE'] = 'prueba'
tryton = Tryton(app, configure_jinja=True)
Prueba = tryton.pool.get('ack.prueba')
Plan = tryton.pool.get('ack.plan')
Employee = tryton.pool.get('ack.employee')

@app.route('/1')
@tryton.transaction()
def hola2():
    registros_prueba = Prueba.search([])
    registros_employee = Employee.search([])
    print(registros_employee)
    return render_template('exportar3.html', lines=registros_prueba, employee=registros_employee)

@app.route('/')
@tryton.transaction()
def hola():

    plan_prueba = Plan.search([])

    lines = []
    for plan in plan_prueba:
        client = {'cliente': plan.cliente, 'tecnico': []}
        technician_names = plan.tecnico.split(',')

        for technician in technician_names:
            # employee = {'tecnico': technician, 'enlaces': []}
            # plan_names = plan.nombre.split(',')
            employee = {
                'tecnico': technician,
                'enlaces': [],
                'from_date': plan.from_date,
                'from_hour': plan.from_hour,
                'to_date': plan.to_date,
                'to_hour': plan.to_hour,
                'tipo': plan.tipo,
                'time': plan.tiempo,
                'prioridad': plan.prioridad,
                'dedicacion': plan.dedicacion,
                'estado': plan.estado
            }
            plan_names = plan.nombre.split(',')
            for allocation in plan_names:
                employee['enlaces'].append({'id': plan.id, 'nombre': allocation})
            client['tecnico'].append(employee)

        lines.append(client)
        print(lines)

    return render_template('exportar3.html', lines=lines)





# @app.route('/api/calendar/events/<record("company.employee"):employee>')
@app.route('/events')
@tryton.transaction()
# @login_required
def calendar_duplicated():
    prueba_id = request.args.get('event_param', type=int)

    if prueba_id:
        recoger_id = Prueba.browse([prueba_id])
        all_tecnicos = Prueba.search([])
        return render_template('event2_old.html', prueba=recoger_id, tecnicos= all_tecnicos)
    else:
        return 'algo salio mal.....'



@app.route('/ruta-flask', methods=['POST'])
@tryton.transaction()
def procesar_duplicacion():
    def obtener_datos_duplicados(data):

        return {key: f"{value}" for key, value in data.items()}
    try:
        data = request.get_json()
        duplicados = obtener_datos_duplicados(data)

        #pdb.set_trace()
        if duplicados is not None:
            plan = Plan()

            plan.cliente = duplicados.get('cliente', '')
            plan.tecnico = duplicados.get('employee', '')
            plan.nombre = duplicados.get('estado_tarea', '').split()[1] if 'estado_tarea' in duplicados else ''
            # Manejar fechas y horas correctamente
            from_date_parts = duplicados.get('from_date', '').split()
            to_date_parts = duplicados.get('to_date', '').split()
            plan.from_date = from_date_parts[0] if len(from_date_parts) > 0 else ''
            from_hour = from_date_parts[1].replace('-', ':') if len(from_date_parts) > 1 else ''
            plan.from_hour = datetime.strptime(from_hour, '%H:%M').time()
            plan.to_date = to_date_parts[0] if len(to_date_parts) > 0 else ''
            to_hour = to_date_parts[1].replace('-', ':') if len(to_date_parts) > 1 else ''
            plan.to_hour = datetime.strptime(to_hour, '%H:%M').time()
            plan.tipo = duplicados.get('estado_tarea', '').split()[0].strip(
                "[]") if 'estado_tarea' in duplicados else ''
            plan.prioridad = duplicados.get('priority', '')
            tiempostr = duplicados.get('time', '')
            tiempostr = tiempostr.replace('-', ':') if len(from_date_parts) > 1 else ''
            tiempohr = datetime.strptime(tiempostr, '%H:%M').time()
            time_delta = timedelta(hours=tiempohr.hour, minutes=tiempohr.minute)
            plan.tiempo = time_delta
            plan.dedicacion = '-'
            plan.estado = duplicados.get('estado', '')
            print(plan)
            plan.save()

            # enviamos todos los registros....
            registros_prueba = Prueba.search([])
            return render_template("plantilla.html", datos_procesados=registros_prueba)
        else:
            print("No se ha podido guardar")
        return redirect("exportar", )
    except Exception as e:
        print('Error en la función procesar_duplicacion:', str(e))
        return jsonify({'error': 'Ocurrió un error en el servidor'}), 500





# logica para exportar a excel

@app.route('/exportar', methods=['POST'])
@tryton.transaction()
def exportar():
    if request.method == 'POST':
        datos_checkbox = request.form.getlist('checkboxes')
        pasar =[]
        # recoger datos
        for dato in datos_checkbox:
            prueba = Prueba.search([('id', '=', dato)])
            if prueba:
                pasar.append(prueba)

        return render_template('excelExport.html', values=pasar)

@app.route('/excelExport', methods=['POST'])
def excelExport():
    # Obtener los datos del campo oculto
    # Obtener los datos del campo oculto
    datos_json = request.form.get('data')

    def agregar_tildes(cadena):
        tildes = {'dedicacion': 'Dedicación', 'tecnico': 'Técnico',
                  'descripcion': 'Descripción'}  # Agrega más palabras según sea necesario
        palabras = cadena.split()
        resultado = ' '.join(tildes.get(palabra.lower(), palabra) for palabra in palabras)
        return resultado



    try:
        # Cargar datos JSON

        datos = json.loads(datos_json)
    except json.decoder.JSONDecodeError as e:
        print("Error al decodificar JSON:", str(e))
        return "Error al decodificar JSON"

        print(datos)
    dataframes = []
    diccionario_combinado = {}
    # Asegúrate de que hay datos y al menos un diccionario en la lista
    if datos and isinstance(datos[0], list):
        # Iterar sobre cada conjunto de datos en la lista
        for conjunto_datos in datos:
            # Crear un diccionario combinado para el conjunto actual
            diccionario_combinado = {}
            for diccionario in conjunto_datos:
                diccionario_combinado.update(diccionario)

            # Crear DataFrame de Pandas con el conjunto actual de datos
            df = pd.DataFrame([diccionario_combinado])

            # Aplicar la función agregar_tildes al nombrar las columnas del DataFrame
            df.columns = df.columns.map(agregar_tildes)

            # Agregar el DataFrame actual a la lista
            dataframes.append(df)

        # Concatenar todos los DataFrames en uno solo
        df_final = pd.concat(dataframes, ignore_index=True)

        # Reemplazar caracteres con tilde en las cabeceras
        df_final = df_final.rename(columns=lambda x: agregar_tildes(x))

        # Crear un nuevo libro de trabajo de Excel con openpyxl
        libro = Workbook()

        # Seleccionar la hoja de trabajo activa
        hoja = libro.active

        # Configurar el estilo para el encabezado (negrita y centrado)
        estilo_encabezado = Font(bold=True)
        alineacion_centro = Alignment(horizontal='center', vertical='center')

        # Configurar el estilo para el color de fondo de la cabecera
        fill_cabecera = PatternFill(start_color="223776", end_color="223776", fill_type="solid")

        # Configurar el estilo para el color de fondo de la tabla
        fill_tabla = PatternFill(start_color="b3d7ff", end_color="b3d7ff", fill_type="solid")

        # Después de insertar los encabezados en la hoja de trabajo
        for columna, encabezado in enumerate(df_final.columns, start=1):
            celda = hoja.cell(row=1, column=columna, value=encabezado)
            celda.font = estilo_encabezado
            celda.alignment = alineacion_centro


            # Ajustar el ancho de la columna según la longitud máxima del encabezado o los datos en esa columna
            longitud_maxima = max(len(str(encabezado)), df_final[encabezado].astype(str).apply(len).max()) + 2
            hoja.column_dimensions[hoja.cell(row=1, column=columna).column_letter].width = longitud_maxima

            celda.fill = fill_cabecera  # Color de fondo para la cabecera

            # Configurar bordes para cada celda
            bordes = Border(left=Side(border_style='thin'), right=Side(border_style='thin'),
                            top=Side(border_style='thin'), bottom=Side(border_style='thin'))
            celda.border = bordes

        # Después de insertar los datos en la hoja de trabajo
        for fila, valores in enumerate(df_final.values, start=2):
            for columna, valor in enumerate(valores, start=1):
                celda = hoja.cell(row=fila, column=columna, value=valor)
                celda.fill = fill_tabla  # Color de fondo para la tabla

                # Configurar bordes para cada celda
                bordes = Border(left=Side(border_style='thin'), right=Side(border_style='thin'),
                                top=Side(border_style='thin'), bottom=Side(border_style='thin'))
                celda.border = bordes

                # Justificar a la derecha
                celda.alignment = Alignment(horizontal='right')

        # Guardar el libro de trabajo en un archivo Excel
        libro.save('datos_formato.xlsx')

        # Devolver el archivo Excel como respuesta
        return send_file('datos_formato.xlsx', as_attachment=True)
    else:
        return "Datos no válidos"
if __name__ == '__main__':
    app.run(debug=True)
