from flask import Flask, render_template, request, send_file
from flask_tryton import Tryton
import json
from flask_cors import CORS
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side



app = Flask(__name__)
CORS(app)
app.config['TRYTON_DATABASE'] = 'prueba'
tryton = Tryton(app, configure_jinja=True)
Prueba = tryton.pool.get('ack.prueba')


@app.route('/')
@tryton.transaction()
def hola():
    prueba = Prueba.search([])
    return render_template('exportar2.html', lines=prueba)


@app.route('/exportar', methods=['POST'])
@tryton.transaction()
def exportar():
    if request.method == 'POST':
        datos_checkbox = request.form.getlist('checkboxes')
        pasar =[]
        # recoger datos
        for dato in datos_checkbox:
            prueba = Prueba.search([('id', '=', dato)])
            if prueba:
                pasar.append(prueba)

        return render_template('exportar_excel.html', prueba=pasar)
    return render_template('exportar2.html')

@app.route('/exportar_excel', methods=['POST'])
def exportar_excel():
    # Obtener los datos del campo oculto
    # Obtener los datos del campo oculto
    datos_json = request.form.get('datos')

    def agregar_tildes(cadena):
        tildes = {'dedicacion': 'Dedicación', 'tecnico': 'Técnico',
                  'descripcion': 'Descripción'}  # Agrega más palabras según sea necesario
        palabras = cadena.split()
        resultado = ' '.join(tildes.get(palabra.lower(), palabra) for palabra in palabras)
        return resultado

    try:

        # Cargar datos JSON
        datos = json.loads(datos_json)
    except json.decoder.JSONDecodeError as e:
        print("Error al decodificar JSON:", str(e))
        return "Error al decodificar JSON"
    print(datos)
    dataframes = []
    diccionario_combinado = {}
    # Asegúrate de que hay datos y al menos un diccionario en la lista
    if datos and isinstance(datos[0], list):
        # Iterar sobre cada conjunto de datos en la lista
        for conjunto_datos in datos:
            # Crear un diccionario combinado para el conjunto actual
            diccionario_combinado = {}
            for diccionario in conjunto_datos:
                diccionario_combinado.update(diccionario)

            # Crear DataFrame de Pandas con el conjunto actual de datos
            df = pd.DataFrame([diccionario_combinado])

            # Aplicar la función agregar_tildes al nombrar las columnas del DataFrame
            df.columns = df.columns.map(agregar_tildes)

            # Agregar el DataFrame actual a la lista
            dataframes.append(df)

        # Concatenar todos los DataFrames en uno solo
        df_final = pd.concat(dataframes, ignore_index=True)

        # Reemplazar caracteres con tilde en las cabeceras
        df_final = df_final.rename(columns=lambda x: agregar_tildes(x))

        # Crear un nuevo libro de trabajo de Excel con openpyxl
        libro = Workbook()

        # Seleccionar la hoja de trabajo activa
        hoja = libro.active

        # Configurar el estilo para el encabezado (negrita y centrado)
        estilo_encabezado = Font(bold=True)
        alineacion_centro = Alignment(horizontal='center', vertical='center')

        # Configurar el estilo para el color de fondo de la cabecera
        fill_cabecera = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

        # Configurar el estilo para el color de fondo de la tabla
        fill_tabla = PatternFill(start_color="EEEEEE", end_color="EEEEEE", fill_type="solid")

        # Después de insertar los encabezados en la hoja de trabajo
        for columna, encabezado in enumerate(df_final.columns, start=1):
            celda = hoja.cell(row=1, column=columna, value=encabezado)
            celda.font = estilo_encabezado
            celda.alignment = alineacion_centro

            # Ajustar el ancho de la columna según la longitud máxima del encabezado o los datos en esa columna
            longitud_maxima = max(len(str(encabezado)), df_final[encabezado].astype(str).apply(len).max()) + 2
            hoja.column_dimensions[hoja.cell(row=1, column=columna).column_letter].width = longitud_maxima

            celda.fill = fill_cabecera  # Color de fondo para la cabecera

            # Configurar bordes para cada celda
            bordes = Border(left=Side(border_style='thin'), right=Side(border_style='thin'),
                            top=Side(border_style='thin'), bottom=Side(border_style='thin'))
            celda.border = bordes

        # Después de insertar los datos en la hoja de trabajo
        for fila, valores in enumerate(df_final.values, start=2):
            for columna, valor in enumerate(valores, start=1):
                celda = hoja.cell(row=fila, column=columna, value=valor)
                celda.fill = fill_tabla  # Color de fondo para la tabla

                # Configurar bordes para cada celda
                bordes = Border(left=Side(border_style='thin'), right=Side(border_style='thin'),
                                top=Side(border_style='thin'), bottom=Side(border_style='thin'))
                celda.border = bordes

                # Justificar a la derecha
                celda.alignment = Alignment(horizontal='right')

        # Guardar el libro de trabajo en un archivo Excel
        libro.save('datos_formato.xlsx')

        # Devolver el archivo Excel como respuesta
        return send_file('datos_formato.xlsx', as_attachment=True)
    else:
        return "Datos no válidos"
if __name__ == '__main__':
    app.run(debug=True)
